import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from collections import defaultdict
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import time
from PIL import Image
from io import BytesIO
import re

def take_screenshot(url):
    # Встановлюємо сервіс для драйвера
    service = Service(ChromeDriverManager().install())

    # Налаштовуємо браузер
    options = Options()
    options.add_argument('--headless')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-gpu')
    options.add_argument('window-size=1920x1080')
    options.add_argument('--remote-debugging-port=9222')  # Додаємо порт для налагодження

    # Запускаємо браузер
    driver = webdriver.Chrome(service=service, options=options)
    
    # Відкриваємо сторінку за URL
    try:
        driver.get(url)
        time.sleep(2)  # даємо час на завантаження сторінки
    except Exception as e:
        print(f"Failed to load URL {url}: {e}")
        driver.quit()
        return None

    # Робимо скріншот всієї сторінки
    screenshot = driver.get_screenshot_as_png()

    # Закриваємо браузер
    driver.quit()

    # Відкриваємо скріншот з використанням PIL
    image = Image.open(BytesIO(screenshot))
    return image

def extract_url(text):
    match = re.search(r'"(.*?)"', text)
    return match.group(1) if match else None

def top_3_by_brand(file_path, sheet_name, new_sheet_name='Top3ByBrand'):
    # Завантажуємо файл
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    # Збираємо всі дані з листа
    data = list(sheet.iter_rows(values_only=True))

    # Отримуємо індекси потрібних колонок
    headers = data[0]
    required_columns = ['Бренд', 'ppr', 'Место публикации', 'URL']

    # Перевіряємо, чи всі необхідні колонки є в заголовках
    for col in required_columns:
        if col not in headers:
            raise ValueError(f"Column '{col}' not found in sheet '{sheet_name}'")

    brand_index = headers.index('Бренд')
    ppr_index = headers.index('ppr')
    place_index = headers.index('Место публикации')
    url_index = headers.index('URL')

    # Створюємо словник для збереження даних по брендам
    brand_dict = defaultdict(list)

    # Наповнюємо словник даними
    for row in data[1:]:
        brand = row[brand_index]
        ppr = row[ppr_index]
        place = row[place_index]
        url = row[url_index]
        #url = extract_url(row[url_index])
        #if url:
        #    brand_dict[brand].append((place, ppr, url))
        brand_dict[brand].append((place, ppr, url))

   # Створюємо новий лист для результатів
    if new_sheet_name in workbook.sheetnames:
        workbook.remove(workbook[new_sheet_name])
    new_sheet = workbook.create_sheet(new_sheet_name)

    # Записуємо заголовки в новий лист
    new_sheet.append(['Бренд', 'Место публикации', 'ppr', 'url'])

    # Виводимо топ-3 по кожному бренду і записуємо в новий лист
    for brand, entries in brand_dict.items():
        # Сортуємо записи по ppr по спаданню
        sorted_entries = sorted(entries, key=lambda x: x[1], reverse=True)
        top_3 = sorted_entries[:3]
        
        for entry in top_3:
            place, ppr, url = entry
            #new_sheet.append([brand, place, ppr, url])
            new_row = [brand, place, ppr, url]
          
            # Робимо скріншот і зберігаємо його
            """screenshot = take_screenshot(url)
            if screenshot is not None:
                screenshot_path = f'{brand}_{place}_{ppr}.png'
                screenshot.save(screenshot_path)
                
                # Вставляємо зображення в новий лист
                img = OpenpyxlImage(screenshot_path)
                new_row.append(screenshot_path)
                new_sheet.append(new_row)
                img_cell = new_sheet.cell(row=new_sheet.max_row, column=new_sheet.max_column)
                img.anchor = img_cell.coordinate
                new_sheet.add_image(img)
            else:
                new_row.append('Failed to load')
                new_sheet.append(new_row) """
            new_sheet.append(new_row)
    # Зберігаємо зміни в файлі
    workbook.save(file_path)
   

def hide_columns_by_names(file_path, sheet_name, columns_to_hide):
    # Завантажуємо файл
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    # Перетворюємо строки колонок у список
    columns_to_hide_list = columns_to_hide.split(',')

    # Проходимо по кожній колонці в першому рядку
    for col in sheet.iter_cols(1, sheet.max_column):
        col_letter = col[0].column_letter
        col_name = col[0].value

        # Якщо назва колонки є в списку колонок, які потрібно сховати
        if col_name in columns_to_hide_list:
            sheet.column_dimensions[col_letter].hidden = True

    # Зберігаємо зміни в файлі
    workbook.save(file_path)

def sort_excel_by_column(file_path, sheet_name, column_name):
    # Завантажуємо файл
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    # Отримуємо індекс колонки за її назвою
    col_index = None
    for col in sheet.iter_cols(1, sheet.max_column):
        if col[0].value == column_name:
            col_index = col[0].col_idx
            break

    if col_index is None:
        raise ValueError(f"Column '{column_name}' not found in sheet '{sheet_name}'")

    # Збираємо всі дані з листа
    data = list(sheet.iter_rows(values_only=True))

    # Зберігаємо заголовки
    headers = data[0]
  
    # Сортуємо дані (окрім заголовків) за вказаною колонкою по спаданню
    sorted_data = sorted(data[1:], key=lambda x: x[col_index-1], reverse=True)

    # Записуємо відсортовані дані назад в лист
    for row_idx, row in enumerate([headers] + sorted_data, 1):
        for col_idx, value in enumerate(row, 1):
            sheet.cell(row=row_idx, column=col_idx, value=value)

    # Зберігаємо зміни в файлі
    workbook.save(file_path)
# Приклад використання функції
file_path = r'c:\d\work\analityka\Дарниця см 17-23.08.xlsx'
sheet_name = 'реєстр'
columns_to_hide = """Профиль,Подписчики,Демография,Возраст,Источник,Тип поста,Время,Сохранено,Заголовок,Профиль,
            места публикации,Аспекты,Тематика,Автокатегория,Потенциальный охват,
            Тип источника,Страна,Регион,Город,Заметки,Сумма всех реакций,Вовлечение,Лайки,Love,Haha,Wow,Sad,Angry,Care	,Dislikes,
            Комментарии,Репосты,Просмотры,Рейтинг,Объекты на изображении,Назначено,
            Профиль места публикации,Подписчики места публикации,Тип источника,Care,Комментарии,Назначено,
            Профиль места публикации
            """
hide_columns_by_names(file_path, sheet_name, columns_to_hide)
print('Колонки сховано')
column_name = 'ppr' # Назва колонки для сортування
sort_excel_by_column(file_path, sheet_name, column_name)
print('Файл відсортовано')
top_3_by_brand(file_path, sheet_name)
print('Топ-3 готово')